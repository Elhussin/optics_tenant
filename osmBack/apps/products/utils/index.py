from django.core.mail import send_mail
from django.conf import settings
from django.template.loader import render_to_string
from django.utils import timezone

import hashlib


LENS_POWERS = [(f"{x:+.2f}", f"{x:+.2f}") for x in [i / 4 for i in range(-120, 121)]]

spherical_lens_powersMunis = [(f"-{abs(x):05.2f}", f"-{abs(x):05.2f}") for x in [i / 4 for i in range(-120, 1)]]
spherical_lens_powersPlus = [(f"+{abs(x):05.2f}", f"+{abs(x):05.2f}") for x in [i / 4 for i in range(1, 121)]]
cylinder_lens_powers = [(f"-{abs(x):05.2f}", f"-{abs(x):05.2f}") for x in [i / 4 for i in range(-60, 1)]]
additional_lens_powers = [(f"+{abs(x):05.2f}", f"+{abs(x):05.2f}") for x in [i / 4 for i in range(1, 25)]]

# دمج الموجب والسالب للـ spherical
spherical_lens_powers = spherical_lens_powersMunis + spherical_lens_powersPlus



def format_decimal(value, width=6, decimals=2):
    if value is None:
        return "—"
    return f"{(value):+0{width}.{decimals}f}"


def send_low_stock_email(stock_items, recipient_email=None):
    """إرسال إيميل للمخزون المنخفض"""
    if not recipient_email:
        recipient_email = settings.DEFAULT_FROM_EMAIL
    
    subject = "تنبيه: مخزون منخفض"
    
    html_message = render_to_string('emails/low_stock_alert.html', {
        'stock_items': stock_items,
        'date': timezone.now().date(),
    })
    
    send_mail(
        subject=subject,
        message="",
        from_email=settings.DEFAULT_FROM_EMAIL,
        recipient_list=[recipient_email],
        html_message=html_message,
    )


def generate_barcode(product_variant):
    """إنشاء باركود للمنتج"""
    import barcode
    from barcode.writer import ImageWriter
    from io import BytesIO
    
    code = barcode.get('code128', str(product_variant.id), writer=ImageWriter())
    buffer = BytesIO()
    code.write(buffer)
    buffer.seek(0)
    
    return buffer.getvalue()


def export_stock_to_excel(stock_queryset):
    """تصدير المخزون إلى Excel"""
    import openpyxl
    from openpyxl.utils import get_column_letter
    from django.http import HttpResponse
    
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = "Stock Report"
    
    # Headers
    headers = [
        'Branch', 'Product', 'Variant', 'Stock Quantity', 
        'Reserved', 'Available', 'Min Level', 'Max Level', 
        'Average Cost', 'Last Cost', 'Last Updated'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = worksheet.cell(row=1, column=col)
        cell.value = header
        cell.font = openpyxl.styles.Font(bold=True)
    
    # Data
    for row, stock in enumerate(stock_queryset, 2):
        worksheet.cell(row=row, column=1).value = str(stock.branch)
        worksheet.cell(row=row, column=2).value = stock.variant.product.name
        worksheet.cell(row=row, column=3).value = str(stock.variant)
        worksheet.cell(row=row, column=4).value = stock.stock_quantity
        worksheet.cell(row=row, column=5).value = stock.quantity_reserved
        worksheet.cell(row=row, column=6).value = stock.available_quantity
        worksheet.cell(row=row, column=7).value = stock.minimum_stock_level
        worksheet.cell(row=row, column=8).value = stock.max_stock_level
        worksheet.cell(row=row, column=9).value = float(stock.average_cost)
        worksheet.cell(row=row, column=10).value = float(stock.last_cost)
        worksheet.cell(row=row, column=11).value = stock.last_updated.strftime('%Y-%m-%d %H:%M')
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        worksheet.column_dimensions[get_column_letter(col)].auto_size = True
    
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=stock_report.xlsx'
    
    workbook.save(response)
    return response



